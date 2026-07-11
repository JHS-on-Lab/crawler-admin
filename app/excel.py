"""엑셀(.xlsx) 다운로드 공통 모듈.

화면 리스트 조회 결과를 xlsx로 내려받을 때는 이 모듈만 재사용한다.
화면(라우트)마다 다른 건 컬럼 정의(ExcelColumn 목록)뿐이고,
워크북 생성·다운로드 응답 헤더 구성은 여기서 공통 처리한다.

사용 예:
    from app.excel import ExcelColumn, xlsx_response

    columns = [
        ExcelColumn("host", "도메인"),
        ExcelColumn("rules_enabled", "규칙", formatter=lambda v: "활성" if v else "비활성"),
    ]
    return xlsx_response(rows, columns, filename="도메인_규칙")
"""

from __future__ import annotations

from collections.abc import Callable, Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any
from urllib.parse import quote

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


@dataclass
class ExcelColumn:
    """엑셀 한 열(column)의 정의.

    key       — row(dict-like)에서 값을 꺼낼 키.
    header    — 시트 1행에 표시할 헤더명.
    formatter — 값 변환 함수(선택). 예: bool → "활성"/"비활성", 비율 → 반올림.
                None 이면 원본값을 그대로 쓴다(datetime/date 는 자동 처리).
    width     — 열 너비(선택). 미지정 시 헤더 길이 기준으로 자동 산정.
    """
    key: str
    header: str
    formatter: Callable[[Any], Any] | None = None
    width: int | None = None


# 수식/CSV 인젝션 방어 — 이 문자로 시작하는 문자열 값은 그대로 셀에 넣으면 안 된다.
# 특히 openpyxl 은 "=" 로 시작하는 문자열을 자동으로 실제 수식 셀(data_type='f')로
# 저장하므로(직접 확인함), 관리자가 이 xlsx를 열면 그 즉시 수식이 실행된다 —
# 이론적 위협이 아니라 실제로 재현되는 문제다.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _neutralize_formula(value: str) -> str:
    if value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def _cell_value(row: Mapping, col: ExcelColumn) -> Any:
    value = row.get(col.key)
    if col.formatter is not None:
        value = col.formatter(value)
    elif isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value  # openpyxl 은 tz-aware datetime 미지원
    elif isinstance(value, (dict, list)):
        value = str(value)
    if isinstance(value, str):
        return _neutralize_formula(value)
    return value


def build_workbook(
    rows: Iterable[Mapping],
    columns: list[ExcelColumn],
    sheet_name: str = "Sheet1",
) -> Workbook:
    """rows 를 columns 정의에 따라 openpyxl Workbook 으로 변환한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # 엑셀 시트명 31자 제한

    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col.header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(i)].width = col.width or max(12, len(col.header) + 4)

    for r, row in enumerate(rows, start=2):
        for c, col in enumerate(columns, start=1):
            ws.cell(row=r, column=c, value=_cell_value(row, col))

    ws.freeze_panes = "A2"
    return wb


def xlsx_response(
    rows: Iterable[Mapping],
    columns: list[ExcelColumn],
    filename: str,
    sheet_name: str = "Sheet1",
) -> Response:
    """rows 를 xlsx 로 변환해 다운로드용 Response 로 반환한다.

    filename: 확장자(.xlsx) 없이 전달. 한글 파일명도 안전하게 인코딩된다.
    """
    wb = build_workbook(rows, columns, sheet_name=sheet_name)
    buf = BytesIO()
    wb.save(buf)

    # 구형 클라이언트는 filename=(ASCII) 쪽만 읽으므로, 한글만 있는 파일명처럼
    # ASCII로 남는 게 없거나 의미 없는 경우 일반 영문 이름으로 대체한다.
    ascii_stripped = filename.encode("ascii", errors="ignore").decode()
    ascii_fallback = ascii_stripped if any(c.isalnum() for c in ascii_stripped) else "export"
    encoded_name = quote(f"{filename}.xlsx")
    return Response(
        content=buf.getvalue(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_fallback}.xlsx"; '
                f"filename*=UTF-8''{encoded_name}"
            ),
        },
    )
